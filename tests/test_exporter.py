from io import BytesIO
import unittest

import pandas as pd
from openpyxl import Workbook, load_workbook

from src.exporter import export_mercado_libre_excel


class MercadoLibreExporterTest(unittest.TestCase):
    def build_template(self):
        wb = Workbook()
        ayuda = wb.active
        ayuda.title = "Ayuda"
        ayuda["A1"] = "contenido ayuda intacto"
        hidden = wb.create_sheet("hidden")
        hidden["A1"] = "contenido hidden intacto"
        promos = wb.create_sheet("Promociones")
        promos.append(["Instrucción general", None, None, None, None, None, None, "NO TOCAR"])
        promos.append(["ITEM_ID", "SKU", "TITLE", "ORIGINAL_PRICE", "DISCOUNT_PERCENTAGE", "FINAL_PRICE", "ACTION", "STATUS"])
        promos.append(["MLA1", "SKU1", "Producto 1", 1000, 5, 950, "No participar", "ok"])
        promos.append(["MLA2", "SKU2", "Producto 2", 2000, 10, 1800, "Participar", "ok"])
        promos.append([None, "SKU3", "Producto 3", 3000, 15, 2550, "Participar", "ok"])
        promos["E3"].number_format = "0%"
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_export_preserves_sheets_columns_instruction_rows_and_only_updates_allowed_cells(self):
        template = self.build_template()
        simulated = pd.DataFrame(
            [
                {
                    "_ROW_ID": "ITEM_ID::MLA1",
                    "ITEM_ID": "MLA1",
                    "SKU": "SKU1",
                    "TITLE": "Producto 1 editado internamente",
                    "DISCOUNT_PERCENTAGE": 20,
                    "FINAL_PRICE": 800,
                    "ACTION": "Participar",
                    "Modificado": "Sí",
                    "Costo": 100,
                },
                {
                    "_ROW_ID": "ITEM_ID::MLA2",
                    "ITEM_ID": "MLA2",
                    "SKU": "SKU2",
                    "TITLE": "Producto 2",
                    "DISCOUNT_PERCENTAGE": 10,
                    "FINAL_PRICE": 1800,
                    "ACTION": "No participar",
                    "Modificado": "No",
                },
                {
                    "_ROW_ID": "SKU_TITLE::SKU3::Producto 3",
                    "SKU": "SKU3",
                    "TITLE": "Producto 3",
                    "DISCOUNT_PERCENTAGE": 25,
                    "FINAL_PRICE": 2250,
                    "ACTION": "Participar",
                    "Modificado": "Sí",
                },
            ]
        )

        exported = export_mercado_libre_excel(template, simulated)
        wb = load_workbook(BytesIO(exported))
        promos = wb["Promociones"]

        self.assertIn("Ayuda", wb.sheetnames)
        self.assertIn("hidden", wb.sheetnames)
        self.assertIn("Promociones", wb.sheetnames)
        self.assertEqual(wb["Ayuda"]["A1"].value, "contenido ayuda intacto")
        self.assertEqual(wb["hidden"]["A1"].value, "contenido hidden intacto")
        self.assertEqual(promos.max_column, 8)
        self.assertEqual([cell.value for cell in promos[1]], ["Instrucción general", None, None, None, None, None, None, "NO TOCAR"])

        self.assertEqual(promos["A3"].value, "MLA1")
        self.assertEqual(promos["B3"].value, "SKU1")
        self.assertEqual(promos["C3"].value, "Producto 1")
        self.assertEqual(promos["D3"].value, 1000)
        self.assertEqual(promos["H3"].value, "ok")
        self.assertEqual(promos["E3"].value, 20)
        self.assertEqual(promos["F3"].value, 800)
        self.assertEqual(promos["G3"].value, "Participar")

        self.assertEqual(promos["E4"].value, 10)
        self.assertEqual(promos["F4"].value, 1800)
        self.assertEqual(promos["G4"].value, "No participar")

        self.assertEqual(promos["E5"].value, 25)
        self.assertEqual(promos["F5"].value, 2250)
        self.assertEqual(promos["G5"].value, "Participar")


class InternalReportExporterTest(unittest.TestCase):
    def test_internal_report_includes_enriched_rows_summary_sorting_and_formats(self):
        from src.exporter import INTERNAL_REPORT_COLUMNS, export_internal_report_excel

        simulated = pd.DataFrame(
            [
                {
                    "SKU": "SKU2",
                    "ITEM_ID": "MLA2",
                    "TITLE": "Producto no modificado",
                    "Nombre": "Nombre 2",
                    "Categorías": "Bazar",
                    "Marca": "Marca B",
                    "Código de barras / EAN": "7791234567890.0",
                    "Costo": 1000,
                    "ORIGINAL_PRICE": 2000,
                    "DISCOUNT_PERCENTAGE": 10,
                    "FINAL_PRICE": 1800,
                    "ACTION": "No participar",
                    "Modificado": "No",
                    "Campo modificado": "",
                    "Margen estimado": 800,
                    "Margen %": 0.4444,
                    "Alerta margen": "",
                    "STATUS": "ok",
                    "ERRORS": "",
                    "_HAS_MATCH": True,
                },
                {
                    "SKU": "SKU1",
                    "ITEM_ID": "MLA1",
                    "TITLE": "Producto modificado",
                    "Nombre": "Nombre 1",
                    "Categorías": "Accesorios",
                    "Marca": "Marca A",
                    "Código de barras / EAN": "7.791234567891E+12",
                    "Costo": 33470,
                    "ORIGINAL_PRICE": 50000,
                    "DISCOUNT_PERCENTAGE": 20,
                    "FINAL_PRICE": 40000,
                    "ACTION": "Participar",
                    "Modificado": "Sí",
                    "Campo modificado": "Descuento",
                    "Margen estimado": 6530,
                    "Margen %": 0.16325,
                    "Alerta margen": "",
                    "STATUS": "ok",
                    "ERRORS": "",
                    "_HAS_MATCH": True,
                },
                {
                    "SKU": "SKU3",
                    "ITEM_ID": "MLA3",
                    "TITLE": "Producto sin cruce",
                    "Código de barras / EAN": "",
                    "Costo": pd.NA,
                    "ORIGINAL_PRICE": 3000,
                    "DISCOUNT_PERCENTAGE": 5,
                    "FINAL_PRICE": 2850,
                    "ACTION": "No participar",
                    "Modificado": "No",
                    "Margen estimado": pd.NA,
                    "Margen %": pd.NA,
                    "Alerta margen": "Margen bajo",
                    "_HAS_MATCH": False,
                },
            ]
        )

        exported = export_internal_report_excel(simulated)
        wb = load_workbook(BytesIO(exported))

        self.assertIn("Reporte", wb.sheetnames)
        self.assertIn("Resumen", wb.sheetnames)

        report = wb["Reporte"]
        headers = [cell.value for cell in report[1]]
        self.assertEqual(headers, INTERNAL_REPORT_COLUMNS)
        rows = list(report.iter_rows(min_row=2, values_only=True))
        modified_index = headers.index("Modificado")
        self.assertEqual([row[modified_index] for row in rows], ["Sí", "No", "No"])
        self.assertEqual({row[modified_index] for row in rows}, {"Sí", "No"})

        ean_index = headers.index("Código de barras / EAN")
        self.assertEqual(rows[0][ean_index], "7791234567891")
        self.assertEqual(rows[1][ean_index], "7791234567890")

        cost_index = headers.index("Costo")
        final_price_index = headers.index("FINAL_PRICE")
        margin_index = headers.index("Margen estimado")
        margin_percentage_index = headers.index("Margen %")
        self.assertEqual(rows[0][cost_index], "$ 33.470,00")
        self.assertEqual(rows[0][final_price_index], "$ 40.000,00")
        self.assertEqual(rows[0][margin_index], "$ 6.530,00")
        self.assertEqual(rows[0][margin_percentage_index], "16,32%")

        summary_values = {row[0].value: row[1].value for row in wb["Resumen"].iter_rows(min_row=2, max_col=2)}
        self.assertEqual(summary_values["Total publicaciones ML"], 3)
        self.assertEqual(summary_values["Total productos cruzados"], 2)
        self.assertEqual(summary_values["Total sin cruce"], 1)
        self.assertEqual(summary_values["Total sin costo"], 1)
        self.assertEqual(summary_values["Total sin EAN"], 1)
        self.assertEqual(summary_values["Total modificados"], 1)
        self.assertEqual(summary_values["Total con margen bajo"], 1)


if __name__ == "__main__":
    unittest.main()
