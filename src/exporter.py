"""Exportaciones de promociones para Mercado Libre y reportes internos."""

from __future__ import annotations

from io import BytesIO
from typing import BinaryIO

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from src.loaders import ML_SHEET_NAME, ML_SKU_ALIASES, _read_upload_bytes, detect_ml_header_row
from src.ml_rows import is_real_ml_publication
from src.transformers import build_row_identifier
from src.utils import format_currency, format_percentage, format_plain_text, parse_optional_number

INTERNAL_REPORT_COLUMNS = [
    "SKU",
    "ITEM_ID",
    "TITLE",
    "Nombre",
    "Categorías",
    "Marca",
    "Código de barras / EAN",
    "Costo",
    "ORIGINAL_PRICE",
    "DISCOUNT_PERCENTAGE",
    "FINAL_PRICE",
    "ACTION",
    "Modificado",
    "Campo modificado",
    "Margen estimado",
    "Margen %",
    "Alerta margen",
    "STATUS",
    "ERRORS",
]
INTERNAL_CURRENCY_COLUMNS = {"Costo", "ORIGINAL_PRICE", "FINAL_PRICE", "Margen estimado"}
INTERNAL_SORT_COLUMNS = ["_MODIFICADO_SORT", "Categorías", "Marca", "SKU"]

EDITABLE_COLUMNS = ("DISCOUNT_PERCENTAGE", "FINAL_PRICE", "ACTION")
PARTICIPATE_ACTION = "Participar"
DO_NOT_PARTICIPATE_ACTION = "No participar"


def _worksheet_to_dataframe(sheet: Worksheet) -> pd.DataFrame:
    """Convierte una hoja openpyxl a DataFrame crudo para reutilizar la detección de encabezado."""
    return pd.DataFrame(sheet.iter_rows(values_only=True))


def _header_map(sheet: Worksheet, header_row_number: int) -> dict[str, int]:
    """Mapea nombres de columnas del header de Mercado Libre a índices 1-based de Excel."""
    headers: dict[str, int] = {}
    for cell in sheet[header_row_number]:
        if cell.value is None:
            continue
        name = str(cell.value).strip()
        if name:
            headers[name] = cell.column

    for alias in ML_SKU_ALIASES:
        if alias in headers and "SKU" not in headers:
            headers["SKU"] = headers[alias]
    return headers


def _row_series(sheet: Worksheet, row_number: int, headers: dict[str, int]) -> pd.Series:
    values = {column: sheet.cell(row=row_number, column=column_index).value for column, column_index in headers.items()}
    return pd.Series(values)


def _normalize_action(value: object) -> str:
    text = "" if pd.isna(value) else str(value).strip()
    return PARTICIPATE_ACTION if text == PARTICIPATE_ACTION else DO_NOT_PARTICIPATE_ACTION


def _normalize_export_value(column: str, value: object) -> object:
    if column == "ACTION":
        return _normalize_action(value)
    number = parse_optional_number(value)
    return value if number is None else number


def _build_export_lookup(simulated_df: pd.DataFrame) -> dict[str, dict[str, object]]:
    """Prepara los valores editables finales indexados por ITEM_ID o SKU+TITLE."""
    lookup: dict[str, dict[str, object]] = {}
    if simulated_df.empty:
        return lookup

    for _, row in simulated_df.iterrows():
        row_id = row.get("_ROW_ID") or build_row_identifier(row)
        values = {column: _normalize_export_value(column, row.get(column)) for column in EDITABLE_COLUMNS}
        lookup[str(row_id)] = values
    return lookup


def export_mercado_libre_excel(original_file: BinaryIO | BytesIO | bytes, simulated_df: pd.DataFrame) -> bytes:
    """
    Genera el Excel final para subir a Mercado Libre usando el archivo original como plantilla.

    Solo modifica DISCOUNT_PERCENTAGE, FINAL_PRICE y ACTION en filas reales de la hoja Promociones.
    Las hojas restantes, las filas instructivas, las columnas, formatos y orden originales se conservan.
    """
    content = original_file if isinstance(original_file, bytes) else _read_upload_bytes(original_file)
    workbook = load_workbook(BytesIO(content))
    if ML_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f'El Excel no contiene la hoja "{ML_SHEET_NAME}".')

    sheet = workbook[ML_SHEET_NAME]
    raw_df = _worksheet_to_dataframe(sheet)
    header_row_number = detect_ml_header_row(raw_df) + 1
    headers = _header_map(sheet, header_row_number)

    missing_editable = [column for column in EDITABLE_COLUMNS if column not in headers]
    if missing_editable:
        raise ValueError("Columnas faltantes para exportar: " + ", ".join(missing_editable))

    lookup = _build_export_lookup(simulated_df)
    for row_number in range(header_row_number + 1, sheet.max_row + 1):
        original_row = _row_series(sheet, row_number, headers)
        if not is_real_ml_publication(original_row):
            continue

        row_id = build_row_identifier(original_row)
        values = lookup.get(row_id, {})
        sheet.cell(row=row_number, column=headers["ACTION"]).value = values.get("ACTION", DO_NOT_PARTICIPATE_ACTION)
        for column in ("DISCOUNT_PERCENTAGE", "FINAL_PRICE"):
            if column in values:
                sheet.cell(row=row_number, column=headers[column]).value = values[column]

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def has_modified_rows(df: pd.DataFrame) -> bool:
    """Indica si la simulación contiene al menos una publicación modificada por el usuario."""
    if "Modificado" not in df.columns:
        return False
    return df["Modificado"].fillna("").astype(str).str.strip().eq("Sí").any()


def _internal_report_value(column: str, value: object) -> object:
    """Normaliza valores para que el reporte interno sea legible y seguro para revisión."""
    if column == "Código de barras / EAN":
        return format_plain_text(value)
    if column in INTERNAL_CURRENCY_COLUMNS:
        return format_currency(value)
    if column == "Margen %":
        return format_percentage(value)
    if pd.isna(value):
        return ""
    return value


def _prepare_internal_report_df(simulated_df: pd.DataFrame) -> pd.DataFrame:
    """Construye la hoja principal del reporte interno con filas modificadas y no modificadas."""
    report = simulated_df.copy()
    for column in INTERNAL_REPORT_COLUMNS:
        if column not in report.columns:
            report[column] = pd.NA

    report["_MODIFICADO_SORT"] = report["Modificado"].fillna("").astype(str).str.strip().eq("Sí").astype(int)
    report = report.sort_values(INTERNAL_SORT_COLUMNS, ascending=[False, True, True, True], na_position="last", kind="stable")
    report = report[INTERNAL_REPORT_COLUMNS].copy()
    for column in INTERNAL_REPORT_COLUMNS:
        report[column] = report[column].map(lambda value, column=column: _internal_report_value(column, value))
    return report.reset_index(drop=True)


def _build_internal_summary(simulated_df: pd.DataFrame) -> list[tuple[str, object]]:
    """Calcula métricas de control para la hoja Resumen."""
    has_match = simulated_df.get("_HAS_MATCH", pd.Series(False, index=simulated_df.index)).fillna(False).astype(bool)
    cost = simulated_df.get("Costo", pd.Series(pd.NA, index=simulated_df.index))
    ean = simulated_df.get("Código de barras / EAN", pd.Series(pd.NA, index=simulated_df.index)).map(format_plain_text)
    modified = simulated_df.get("Modificado", pd.Series("", index=simulated_df.index)).fillna("").astype(str).str.strip().eq("Sí")
    alert = simulated_df.get("Alerta margen", pd.Series("", index=simulated_df.index)).fillna("").astype(str).str.strip()

    return [
        ("Total publicaciones ML", int(len(simulated_df))),
        ("Total productos cruzados", int(has_match.sum())),
        ("Total sin cruce", int((~has_match).sum())),
        ("Total sin costo", int(cost.map(parse_optional_number).isna().sum())),
        ("Total sin EAN", int(ean.eq("").sum())),
        ("Total modificados", int(modified.sum())),
        ("Total con margen negativo", int(alert.eq("Margen negativo").sum())),
        ("Total con margen bajo", int(alert.eq("Margen bajo").sum())),
        ("Fecha/hora de generación", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]


def export_internal_report_excel(simulated_df: pd.DataFrame) -> bytes:
    """Genera el Excel de control interno con hoja Reporte y hoja Resumen."""
    workbook = Workbook()
    report_sheet = workbook.active
    report_sheet.title = "Reporte"

    report_df = _prepare_internal_report_df(simulated_df)
    report_sheet.append(INTERNAL_REPORT_COLUMNS)
    for row in report_df.itertuples(index=False, name=None):
        report_sheet.append(row)

    for cell in report_sheet[1]:
        cell.font = Font(bold=True)
    for column_cells in report_sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 45)
        report_sheet.column_dimensions[column_cells[0].column_letter].width = width

    summary_sheet = workbook.create_sheet("Resumen")
    summary_sheet.append(["Métrica", "Valor"])
    for metric, value in _build_internal_summary(simulated_df):
        summary_sheet.append([metric, value])
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
    summary_sheet.column_dimensions["A"].width = 32
    summary_sheet.column_dimensions["B"].width = 24

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
